import asyncio
import re
import openpyxl
import urllib.parse
import os
import random
from playwright.async_api import async_playwright

MAX_CONCURRENT_REQUESTS = 3
EXCEL_FILE = "Климатрейд.xlsx"
OUTPUT_FILE = "Климатрейд_result.xlsx"
BASE_URL = "https://climate-shop.by"

async def process_model(model_name, context, sem, row_idx, results):
    async with sem:
        print(f"[{row_idx}] Processing: {model_name}")
        clean_model = str(model_name).strip()
        clean_model = re.sub(r'\(.*?\)', '', clean_model) # remove text in parentheses
        if '/' in clean_model:
            clean_model = clean_model.split('/')[0] # take first part before slash
        clean_model = clean_model.strip()
        
        search_url = f"{BASE_URL}/catalog/?q={urllib.parse.quote(clean_model)}"
        
        # Adding a random delay to simulate a human and bypass DDoS protection
        await asyncio.sleep(random.uniform(3.0, 7.0))
        
        page = await context.new_page()
        try:
            # Go to search page
            await page.goto(search_url, wait_until="domcontentloaded", timeout=20000)
            
            # Wait for search results or something to load
            try:
                await page.wait_for_selector(".catalog-block-view__item, .catalog-block__item, .list-item", timeout=5000)
            except Exception:
                pass # Proceed anyway
                
            # Find the actual product link
            product_link = None
            locators = [".catalog-block-view__item a", ".catalog-block__item a", ".list-item a"]
            for loc in locators:
                elements = await page.locator(loc).all()
                for el in elements:
                    href = await el.get_attribute("href")
                    if href and href.startswith("/catalog/") and "?q=" not in href:
                        product_link = BASE_URL + href
                        break
                if product_link:
                    break
                    
            if not product_link:
                results[row_idx] = "Нет ссылки"
                print(f"[{row_idx}] Cannot find product card for {model_name}")
                return
                
            print(f"[{row_idx}] Found link: {product_link}")
            
            # Navigate to product page
            response = await page.goto(product_link, wait_until="domcontentloaded", timeout=20000)
            if response is None or response.status == 404:
                results[row_idx] = "Товар не найден (404)"
                print(f"[{row_idx}] 404 for {product_link}")
                return
                
            await page.wait_for_timeout(1500) # give JS time to populate article span
            
            # Extract article
            try:
                article = await page.locator(".js-replace-article").get_attribute("data-value", timeout=3000)
                if article:
                    results[row_idx] = article
                    print(f"[{row_idx}] Found article: {article}")
                    return
            except Exception:
                pass
                
            # Alternative locations
            try:
                sku_meta = await page.locator("meta[itemprop='sku']").get_attribute("content", timeout=1000)
                if sku_meta:
                    results[row_idx] = sku_meta
                    print(f"[{row_idx}] Found article via meta: {sku_meta}")
                    return
            except Exception:
                pass
                
            # Last resort: visual article block
            try:
                art_text = await page.locator(".article .value").inner_text(timeout=1000)
                if art_text:
                    results[row_idx] = art_text.strip()
                    print(f"[{row_idx}] Found article via .article .value: {art_text}")
                    return
            except Exception:
                pass
                
            results[row_idx] = "Артикул не найден в DOM"
            print(f"[{row_idx}] Article not found anywhere in DOM")
            
        except Exception as e:
            print(f"[{row_idx}] Exception processing {model_name}: {e}")
            results[row_idx] = "Ошибка загрузки"
        finally:
            await page.close()

async def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"File {EXCEL_FILE} not found!")
        return

    print(f"Loading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    
    headers = [cell.value for cell in sheet[1]]
    model_col_idx = 0
    for i, h in enumerate(headers):
        if h and "Модель" in str(h):
            model_col_idx = i
            break
            
    print(f"Using column index {model_col_idx} for Models.")
    
    article_col_idx = len(headers)
    for i, h in enumerate(headers):
        if h and "Артикул" in str(h):
            article_col_idx = i
            break
            
    if article_col_idx == len(headers):
        sheet.cell(row=1, column=article_col_idx+1, value="Артикул")
        
    results = {}
    sem = asyncio.Semaphore(MAX_CONCURRENT_REQUESTS)
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        # Setup context with user agent to bypass simple blocks
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800}
        )
        
        tasks = []
        count = 0
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            model_name = row[model_col_idx]
            if model_name:
                tasks.append(process_model(model_name, context, sem, row_idx, results))
                count += 1
        
        print(f"Total tasks to run: {len(tasks)}")
        await asyncio.gather(*tasks)
        
        await browser.close()
        
    for row_idx, article in results.items():
        sheet.cell(row=row_idx, column=article_col_idx+1, value=article)
        
    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done!")

if __name__ == "__main__":
    asyncio.run(main())
